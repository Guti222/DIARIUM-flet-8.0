import sqlite3
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Ensure project root (parent of src/) is on sys.path so src package resolves
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

from src.utils.paths import get_db_path


def exportar_libro_diario(libro_id: int, output_file: str | Path | None = None) -> Path:
    """Genera un Excel del libro diario indicado por id.

    Devuelve la ruta del archivo generado. Si no hay datos, lanza ValueError.
    """
    bd_path = get_db_path()
    conn = sqlite3.connect(bd_path)
    df = pd.read_sql_query(
        """
        SELECT 
            a.id_libro_diario AS Libro_Diario,
            ld.nombre_empresa,
            ld.contador,
            a.id_asiento AS AsientoID,
            a.fecha AS Fecha,
            cc.codigo_cuenta AS Código,
            cc.descripcion AS Descripción,
            la.debe AS Debe,
            la.haber AS Haber,
            a.descripcion AS Comentario
        FROM linea_asiento la 
        INNER JOIN asiento a ON la.id_asiento=a.id_asiento
        INNER JOIN cuenta_contable cc ON la.id_cuenta_contable = cc.id_cuenta_contable
        INNER JOIN libro_diario ld ON a.id_libro_diario = ld.id_libro_diario
        WHERE a.id_libro_diario = ?
        ORDER BY a.id_asiento, la.id_linea_asiento
        """,
        conn,
        params=(libro_id,),
    )
    conn.close()

    if df.empty:
        raise ValueError("No se encontraron registros para exportar.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    nombre_empresa = df.iloc[0]["nombre_empresa"]
    contador = df.iloc[0]["contador"]

    def _apply_header(sheet, title_text: str):
        sheet.insert_rows(1)
        sheet.merge_cells("A1:E1")
        cell_title = sheet["A1"]
        cell_title.value = title_text
        cell_title.font = Font(bold=True, size=16)
        cell_title.alignment = Alignment(horizontal="center", vertical="center")
        cell_title.fill = PatternFill("solid", fgColor="0C46E6")
        cell_title.border = thin_border

        headers = ["Fecha", "Código", "Descripción", "Debe", "Haber"]
        for col_idx, h in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_idx, value=h)
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="2F9EE7")
            cell.border = thin_border

    def _apply_column_formats(sheet):
        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 50
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 20

        for cell in sheet["A"]:
            if cell.row != 1 and cell.value not in (None, ""):
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in sheet["B"]:
            if cell.row != 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ["D", "E"]:
            for cell in sheet[col]:
                if cell.row != 1:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    _apply_header(ws, f"Libro Diario: Empresa({nombre_empresa}) (Contador: {contador})")

    fila = 3
    separador_num = 1

    for id_asiento, grupo in df.groupby("AsientoID", sort=False):
        fecha = grupo.iloc[0]["Fecha"]

        ws.cell(row=fila, column=1, value=fecha)
        ws.cell(row=fila, column=3, value=f"-------({separador_num})-------")
        ws.cell(row=fila, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=fila, column=3).alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, 6):
            ws.cell(row=fila, column=col).border = thin_border

        fila += 1
        separador_num += 1

        for _, r in grupo.iterrows():
            ws.cell(row=fila, column=1, value=None)
            ws.cell(row=fila, column=2, value=r["Código"])

            cell_desc = ws.cell(row=fila, column=3, value=r["Descripción"])
            if r["Debe"] and r["Debe"] > 0:
                cell_desc.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell_desc.alignment = Alignment(horizontal="left", vertical="center", indent=2)

            cell_debe = ws.cell(row=fila, column=4, value=r["Debe"])
            cell_debe.alignment = Alignment(horizontal="center", vertical="center")

            cell_haber = ws.cell(row=fila, column=5, value=r["Haber"])
            cell_haber.alignment = Alignment(horizontal="center", vertical="center")

            for col in range(1, 6):
                ws.cell(row=fila, column=col).border = thin_border

            fila += 1

        comentario = grupo.iloc[0]["Comentario"]
        if comentario and comentario.strip() != "":
            cell_coment = ws.cell(row=fila, column=3, value=comentario)
            cell_coment.font = Font(italic=True)
            cell_coment.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, 6):
                ws.cell(row=fila, column=col).border = thin_border
            fila += 1

    _apply_column_formats(ws)

    # --- Hoja Libro Mayor (formato Cuentas T) ---
    conn = sqlite3.connect(bd_path)
    df_mayor = pd.read_sql_query(
        """
        SELECT
            a.id_asiento AS AsientoID,
            cc.id_cuenta_contable AS CuentaID,
            cc.codigo_cuenta AS Código,
            cc.nombre_cuenta AS NombreCuenta,
            la.debe AS Debe,
            la.haber AS Haber
        FROM linea_asiento la
        INNER JOIN asiento a ON la.id_asiento=a.id_asiento
        INNER JOIN cuenta_contable cc ON la.id_cuenta_contable = cc.id_cuenta_contable
        WHERE a.id_libro_diario = ?
        ORDER BY cc.codigo_cuenta, a.id_asiento, la.id_linea_asiento
        """,
        conn,
        params=(libro_id,),
    )
    conn.close()

    ws_mayor = wb.create_sheet("Libro Mayor")

    # Título de hoja
    ws_mayor.merge_cells("A1:E1")
    cell_title = ws_mayor["A1"]
    cell_title.value = f"Libro Mayor (Cuentas T): Empresa({nombre_empresa}) (Contador: {contador})"
    cell_title.font = Font(bold=True, size=16)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    cell_title.fill = PatternFill("solid", fgColor="0C46E6")
    cell_title.border = thin_border

    fila = 3
    for _, grupo in df_mayor.groupby("CuentaID", sort=False):
        codigo = grupo.iloc[0]["Código"]
        nombre = grupo.iloc[0]["NombreCuenta"]

        # Encabezado de cuenta
        ws_mayor.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
        cell_account = ws_mayor.cell(row=fila, column=1, value=f"{codigo} - {nombre}")
        cell_account.font = Font(bold=True, size=13)
        cell_account.alignment = Alignment(horizontal="center", vertical="center")
        cell_account.fill = PatternFill("solid", fgColor="2F9EE7")
        for col in range(1, 6):
            ws_mayor.cell(row=fila, column=col).border = thin_border
        fila += 1

        # Encabezado Debe / Haber
        ws_mayor.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        ws_mayor.merge_cells(start_row=fila, start_column=4, end_row=fila, end_column=5)
        cell_debe_header = ws_mayor.cell(row=fila, column=1, value="Debe")
        cell_haber_header = ws_mayor.cell(row=fila, column=4, value="Haber")
        for cell in (cell_debe_header, cell_haber_header):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="B7D8FF")
        for col in range(1, 6):
            ws_mayor.cell(row=fila, column=col).border = thin_border
        fila += 1

        debe_rows = [(int(r["AsientoID"]), float(r["Debe"] or 0)) for _, r in grupo.iterrows() if (r["Debe"] or 0) > 0]
        haber_rows = [(int(r["AsientoID"]), float(r["Haber"] or 0)) for _, r in grupo.iterrows() if (r["Haber"] or 0) > 0]

        max_len = max(len(debe_rows), len(haber_rows), 1)
        total_debe = 0.0
        total_haber = 0.0
        for i in range(max_len):
            if i < len(debe_rows):
                asiento_id, monto = debe_rows[i]
                ws_mayor.cell(row=fila, column=1, value=asiento_id)
                ws_mayor.cell(row=fila, column=2, value=monto)
                ws_mayor.cell(row=fila, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws_mayor.cell(row=fila, column=2).alignment = Alignment(horizontal="center", vertical="center")
                total_debe += monto
            if i < len(haber_rows):
                asiento_id, monto = haber_rows[i]
                ws_mayor.cell(row=fila, column=4, value=asiento_id)
                ws_mayor.cell(row=fila, column=5, value=monto)
                ws_mayor.cell(row=fila, column=4).alignment = Alignment(horizontal="center", vertical="center")
                ws_mayor.cell(row=fila, column=5).alignment = Alignment(horizontal="center", vertical="center")
                total_haber += monto
            for col in range(1, 6):
                ws_mayor.cell(row=fila, column=col).border = thin_border
            fila += 1

        # Totales
        ws_mayor.cell(row=fila, column=1, value="Total")
        ws_mayor.cell(row=fila, column=2, value=total_debe)
        ws_mayor.cell(row=fila, column=4, value="Total")
        ws_mayor.cell(row=fila, column=5, value=total_haber)
        for col in (1, 2, 4, 5):
            ws_mayor.cell(row=fila, column=col).font = Font(bold=True)
            ws_mayor.cell(row=fila, column=col).alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, 6):
            ws_mayor.cell(row=fila, column=col).border = thin_border
        fila += 2

    ws_mayor.column_dimensions["A"].width = 12
    ws_mayor.column_dimensions["B"].width = 18
    ws_mayor.column_dimensions["C"].width = 4
    ws_mayor.column_dimensions["D"].width = 12
    ws_mayor.column_dimensions["E"].width = 18

    for col in ["B", "E"]:
        for cell in ws_mayor[col]:
            if cell.row != 1 and cell.value not in (None, ""):
                cell.number_format = "#,##0.00"

    # --- Hoja Plan de Cuentas ---
    try:
        conn = sqlite3.connect(bd_path)
        cur = conn.cursor()
        cur.execute(
            "SELECT COALESCE(id_plan_cuenta, 0) FROM libro_diario WHERE id_libro_diario = ?",
            (libro_id,),
        )
        plan_id_row = cur.fetchone()
        plan_id = int(plan_id_row[0]) if plan_id_row else 0

        df_plan = pd.read_sql_query(
            """
            SELECT
                t.numero_cuenta AS TipoCodigo,
                t.nombre_tipo_cuenta AS TipoNombre,
                r.numero_cuenta AS RubroCodigo,
                r.nombre_rubro AS RubroNombre,
                g.numero_cuenta AS GenericoCodigo,
                g.nombre_generico AS GenericoNombre,
                c.codigo_cuenta AS CuentaCodigo,
                c.nombre_cuenta AS CuentaNombre,
                c.descripcion AS CuentaDescripcion
            FROM tipo_cuenta t
            LEFT JOIN rubro r ON r.id_tipo_cuenta = t.id_tipo_cuenta
            LEFT JOIN generico g ON g.id_rubro = r.id_rubro
            LEFT JOIN cuenta_contable c ON c.id_generico = g.id_generico
            WHERE t.id_plan_cuenta = ?
            ORDER BY t.numero_cuenta, r.numero_cuenta, g.numero_cuenta, c.codigo_cuenta
            """,
            conn,
            params=(plan_id,),
        )
    finally:
        try:
            conn.close()
        except Exception:
            pass

    ws_plan = wb.create_sheet("Plan de Cuentas")
    # Header
    headers_plan = [
        "TipoCodigo",
        "TipoNombre",
        "RubroCodigo",
        "RubroNombre",
        "GenericoCodigo",
        "GenericoNombre",
        "CuentaCodigo",
        "CuentaNombre",
        "CuentaDescripcion",
    ]
    for col_idx, h in enumerate(headers_plan, start=1):
        cell = ws_plan.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="B7D8FF")
        cell.border = thin_border

    # Data
    if not df_plan.empty:
        for r_idx, row in enumerate(df_plan.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_plan.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

    # Column widths
    ws_plan.column_dimensions["A"].width = 14
    ws_plan.column_dimensions["B"].width = 28
    ws_plan.column_dimensions["C"].width = 14
    ws_plan.column_dimensions["D"].width = 28
    ws_plan.column_dimensions["E"].width = 16
    ws_plan.column_dimensions["F"].width = 28
    ws_plan.column_dimensions["G"].width = 16
    ws_plan.column_dimensions["H"].width = 32
    ws_plan.column_dimensions["I"].width = 40

    output_path = Path(output_file) if output_file else Path(f"Libro_Diario_{libro_id}.xlsx")
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    try:
        libro = int(sys.argv[1]) if len(sys.argv) > 1 else 1
        output = exportar_libro_diario(libro)
        print("Archivo Excel creado y formateado correctamente:", output)
    except ValueError as exc:
        print(exc)